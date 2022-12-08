
import datetime
import json
import math
import os
import random
from io import BytesIO
from django.utils.dateparse import parse_datetime
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core import serializers
from django.core.files.storage import FileSystemStorage
from django.core.mail import EmailMessage, send_mail
from django.db.models import Case, Count, F, Q, Sum, Value, When,DurationField, ExpressionWrapper,Func,IntegerField
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView, ListView
from reportlab.pdfgen import canvas
import unidecode
from api.automatizacion_tareas import generar_email_aut
from api.models import *
from servidor.settings import BASE_DIR
import urllib
from  django.utils.datastructures import MultiValueDictKeyError
from .forms import (AsignacionTareaForm, CasosForm, EditarFormGestor,
                    datosuserForm, datosuserFormEdit,
                    informacionComplementaria, seguimientoFormulario,
                    userRegister)

global usuario
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from weasyprint import HTML


@login_required
def perfilUsuario(request):
    usuario = DatosUsuario.objects.filter(
        login_id=request.user.id).select_related('login_id')
    if request.method == 'GET':
        usuario = DatosUsuario.objects.filter(
            login_id=request.user.id).select_related('login_id')

    return render(request, 'informacion_paciente.html', {'usuario': usuario})
'''
registro y crud del usuario
'''

@login_required
def editarUser(request, id):
    try:
        persona = DatosUsuario.objects.get(pk=id)
        if request.method == 'GET':
            persona_form = datosuserFormEdit(instance=persona)
        else:
            persona_form = datosuserFormEdit(request.POST, instance=persona)

            if persona_form.is_valid():
                persona_form.save()
                messages.add_message(request, messages.SUCCESS, message='Se ha editado con exito')
                return redirect('perfil')
            else:
                messages.add_message(
                    request, messages.ERROR, message='Vuelva a intetarlo')
    except Exception as e:
        print(e)

    return render(request, 'editarUser.html', {'persona_form': persona_form, 'persona': persona})


class perfilUsuarioRegistro(CreateView):
    model = DatosUsuario
    template_name = 'registro.html'
    form_class = datosuserForm
    second_form_class = userRegister
    success_url = reverse_lazy('login')

    def get_context_data(self, **kwargs):
        context = super(perfilUsuarioRegistro, self).get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(self.request.GET)
            return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST)
        form2 = self.second_form_class(request.POST)
        if form.is_valid() and form2.is_valid():
            solicitud = form.save(commit=False)
            solicitud.login_id = form2.save()
            solicitud.save()
            messages.add_message(request, messages.SUCCESS,
                                 message='Registro exitoso')
            
            return HttpResponseRedirect(self.get_success_url())
        else:
            messages.add_message(request, messages.ERROR,
                                 message='registro fallido vuelva e intentelo')
            return redirect('login')


def validar_grupo(user):
    return user.groups.filter(name__in=['Administrador','Analista', 'Gestor', 'Paciente']).exists()


def logear(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
      
     
        if not user:
            messages.add_message(request, messages.ERROR,
                                 message='contraseña errada o usuario errado')
            return render(request, 'usuarioPorcorreo.html')

        else:
            if validar_grupo(user):
                login(request, user)
                return redirect('perfil')
            else:
                messages.add_message(request, messages.ERROR,
                                     message='El usuario estara activado pronto')
            return redirect('login')

    return render(request, 'usuarioPorcorreo.html')


def log_out(request):
    
    logout(request)
    messages.add_message(request, messages.SUCCESS, f'ha vuelto al inicio')
    return redirect('login')


def perfiluser(request):
    return render(request, 'estructuraSuap.html')
'''
aqui empieza el crud del paciente
'''

@login_required
def registrarCaso(request):
    datos_usuario =  DatosUsuario.objects.get(login_id=request.user.id)
    caso = Casos.objects.filter(id_usuario = datos_usuario.id_cedula).select_related('estado').filter(estado__idestado=Case(When(estado__nombreestado='abierto',then=Value(1)),When(estado__nombreestado='proceso',then=Value(2)))).count()
    today=datetime.date.today().isoformat()
 
    numeroradicado = math.floor(random.random()* 1000)
    forma_persona = CasosForm(request.POST, request.FILES)
    if request.method == 'POST':
        print(request.POST)
        if caso>=1:
               messages.add_message(request, messages.WARNING,message='No puede crear otro caso hasta que este finalice')
               return redirect('caso') 
        else:
            if forma_persona.is_valid():
                forma_persona.save()
                messages.add_message(request, messages.SUCCESS, message='Se ha guardado con exito')
                return redirect('perfil')
    else:
            initial_data = {'id_usuario':datos_usuario.id_cedula,'estado':1,'fecharesgistrocaso': today,'numeroradicado':numeroradicado}
            forma_persona = CasosForm(initial=initial_data)
    return render(request, 'registrarCaso.html', {'forma_persona': forma_persona})


def page_not_found(request,exception):
    return render(request, '404.html')
    
def page_error(request):
    return render(request,'505.html',status=500)
@login_required
def seguimiento(request):
    datos_usuario=DatosUsuario.objects.get(login_id=request.user.id)

    
    casos=Casos.objects.filter(id_usuario= datos_usuario.id_cedula).last()
    
   
    return render(request,'seguimiento.html',{'casos':casos})
@login_required
def historial_casos(request):
    datos_usuario=DatosUsuario.objects.get(login_id=request.user.id)
    
    try:
        casoshistorial = Casos.objects.filter(id_usuario=datos_usuario.id_cedula)
       
        fechafinal = Casos.objects.values('id_caso','fechaatenfinalizado','fecharesgistrocaso').filter(id_usuario= datos_usuario.id_cedula).annotate(duration=TimeStampDiff(F('fechaatenfinalizado') ,F('fechaatenabierto'),output_field=IntegerField()))    

        print(fechafinal)
    except Casos.DoesNotExist:
        messages.add_message(request,messages.ERROR,message='No existe a un caso creado')

    return render(request,'historial.html',{'casoshistorial': casoshistorial,'dic_fecha':fechafinal})
'''
aqui empieza el crud de los casos
'''

@login_required
def gestorcrud(request):
    casos = Casos.objects.all()
    abierto = Casos.objects.all().filter(estado__idestado=1).count()
    proceso=Casos.objects.all().filter(estado__idestado=2).count()
    finalizado=Casos.objects.all().filter(estado__idestado=3).count()
    modal_status='eliminar'
    return render(request,'Gestor/gestorcrud.html',{'casos':casos,'abierto':abierto,'proceso':proceso,'finalizado':finalizado,'modal_status':modal_status})

@login_required
def gestorCrudDelete(request,id):

    casos=Casos.objects.get(pk=id)
    print(casos)
    return render(request,'Gestor/eliminarcaso.html',{'casos':casos})

@login_required
def ajax_eliminar(request):
    id_caso = request.POST.get('id_caso')

    Casos.objects.filter(id_caso=id_caso).update(estado_pendiente='0')
  
    messages.add_message(request, messages.SUCCESS, message='Se ha elimino con exito')
    response={}
    return JsonResponse(response)

@login_required
def editarCrudGestor(request,id):
   try:
       
    
      
        caso=Casos.objects.get(pk=id)
   
        if request.method == 'GET':

                forma_persona = EditarFormGestor(instance=caso)
                info_com=informacionComplementaria(instance=caso.id_comple_info)
                segui=seguimientoFormulario(instance=caso.id_seguimiento)
        else:
               
                forma_persona = EditarFormGestor(request.POST,request.FILES,instance=caso)
                info_com=informacionComplementaria(request.POST,instance=caso.id_comple_info)
                segui=seguimientoFormulario(request.POST,instance=caso.id_seguimiento)
               
                # files=request.FILES.getlist('formula_medica')
                if forma_persona.is_valid() and info_com.is_valid() and segui.is_valid(): 
              
                       
                        forma_persona=forma_persona.save(commit=False)
                        forma_persona.id_comple_info=info_com.save()
                        forma_persona.id_seguimiento=segui.save()
                        generar_email_aut(caso.id_usuario.login_id.email,caso.estado.idestado,caso.id_caso)
                        forma_persona.save()
                        messages.add_message(request, messages.SUCCESS, message='Se ha editado con exito')
                        return redirect('busqueda')
                else:
                 
                 forma_persona = EditarFormGestor()
                 
   except AttributeError as e:
                   print(e)
               
     
   return render(request,'Gestor/editarCasoGestor.html',{'forma_persona':forma_persona,'caso':caso,'info_com':info_com,'segui':segui})

@login_required
def registrarCasoGestor(request):
    # datos_usuario =  DatosUsuario.objects.get(login_id=request.user.id)
  
    
    infocom = InfoComplementaria.objects.all().last()
    segui = Seguimiento.objects.all().last()
    numeroradicado = math.floor(random.random()* 1000)
    forma_persona = EditarFormGestor(request.POST, request.FILES)
    if request.method == 'POST':
            if forma_persona.is_valid():
                forma_persona=forma_persona.save(commit=False)
                caso = Casos.objects.filter(id_usuario = int(request.POST['id_usuario'])).select_related('estado').filter(estado__idestado=Case(When(estado__nombreestado='abierto',then=Value(1)),When(estado__nombreestado='proceso',then=Value(2)))).count()
                if caso>=1:
                    messages.add_message(request, messages.ERROR,message='este usuario que selecciono ya tiene un caso abierto')
                    return redirect('registrarCasoGestor')
                else:
                    forma_persona.save()
                    messages.add_message(request, messages.SUCCESS,message='ha creado un caso')
                    return redirect('busqueda')
            else:
                messages.add_message(request, messages.ERROR,message='Ingrese informacion complementaria y de seguimiento')
    else:
            initial_data = {'estado':1,'fecharesgistrocaso':datetime.datetime.now(),'numeroradicado':numeroradicado,'id_comple_info':infocom,'id_seguimiento':segui}
            forma_persona = EditarFormGestor(initial=initial_data)
    return render(request, 'Gestor/registroCasoGestor.html', {'forma_persona': forma_persona})


@login_required
def informacionComplementarias(request):

    if request.method == 'POST':
        forma_persona = informacionComplementaria(request.POST)
        
        if forma_persona.is_valid():
            forma_persona = forma_persona.save()
            messages.add_message(request,messages.SUCCESS,message='Se aguardo con exito')
            return redirect('registrarCasoGestor')
        
    else:
        messages.add_message(request, messages.ERROR,message='Ingrese informacion complemtaria de nuevo')
        forma_persona=informacionComplementaria()
    return render(request,'Gestor/informacionComplementaria.html',{'forma_persona':forma_persona})
@login_required           
def informacionComplementariasCrear(request):
     

    forma_persona=informacionComplementaria()
    return render(request,'Gestor/informacionComplementariacreareditar.html',{'forma_persona':forma_persona})    
@login_required
def info_co_post_ajax(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
           forma_persona=informacionComplementaria(request.POST)
           if forma_persona.is_valid():
               forma_persona=forma_persona.save()
               messages.add_message(request,messages.SUCCESS,message='Se guardo con exito')
               serializar=serializers.serialize('json',[forma_persona,])
               return JsonResponse({'exito':serializar},status=200)
           else:
               messages.add_message(request, messages.ERROR,message='Ingrese informacion complemtaria de nuevo')
               return JsonResponse({'error':serializar},status=400)
    return JsonResponse({"error": ""}, status=400)
@login_required
def seguimientoGestor(request):
    seguimientoForm=seguimientoFormulario(request.POST) 
    if request.method=='POST':
        seguimientoForm=seguimientoFormulario(request.POST)
        if seguimientoForm.is_valid():
            seguimientoForm= seguimientoForm.save()
            messages.add_message(request,messages.SUCCESS,message='Se aguardo con exito')
            return redirect('registrarCasoGestor')
        else:
            messages.add_message(request, messages.ERROR,message='Ingrese informacion seguimiento de nuevo')
            seguimientoForm=seguimientoFormulario()
    return render(request,'Gestor/seguimientoGestor.html',{'seguimientoForm':seguimientoForm})

@login_required
def seguimientoGestor_creareditar(request):
    seguimientoForm=seguimientoFormulario() 
   
    return render(request,'Gestor/seguimientoGestorcreareditar.html',{'seguimientoForm':seguimientoForm})
@login_required
def segui_co_post_ajax(request):

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
           segui_form=seguimientoFormulario(request.POST)

           if segui_form.is_valid():
           
               segui_form=segui_form.save()
               messages.add_message(request,messages.SUCCESS,message='Se aguardo con exito')
               segui=serializers.serialize('json',[segui_form,])
               return JsonResponse({'exito':segui},status=200)
           else:
               messages.add_message(request, messages.ERROR,message='Ingrese informacion seguimiento de nuevo')
               return JsonResponse({'error':segui},status=400)
    return JsonResponse({"error": ""}, status=400)
@login_required
def editarInfo(request,id):

      try:
        complementaria=InfoComplementaria.objects.get(pk=id)
        if request.method == 'GET':
                forma_persona = informacionComplementaria(instance=complementaria)
        else:
                forma_persona = informacionComplementaria(request.POST,request.FILES,instance=complementaria)
                
                # files=request.FILES.getlist('formula_medica')
                if forma_persona.is_valid(): 
                    forma_persona.save()
                    messages.add_message(request, messages.SUCCESS, message='Se ha editado con exito')
                    return redirect('busqueda')
                else:
                   messages.add_message(request, messages.ERROR, message='LLENE LOS CAMPOS FALTANTES')
      except AttributeError as e:
                   print(e)
               

      return render(request,'Gestor/editarinfoco.html',{'forma_persona':forma_persona,'complementaria':complementaria})
  
@login_required
def editarSegui(request,id):
      try:
        seguimiento=Seguimiento.objects.get(pk=id)
        if request.method == 'GET':
                forma_persona = seguimientoFormulario(instance=seguimiento)
        else:
                forma_persona = seguimientoFormulario(request.POST,request.FILES,instance=seguimiento)
                
                # files=request.FILES.getlist('formula_medica')
                if forma_persona.is_valid(): 
                    forma_persona.save()
                    messages.add_message(request, messages.SUCCESS, message='Se ha editado con exito')
                    return redirect('busqueda')
                else:
                   messages.add_message(request, messages.ERROR, message='LLENE LOS CAMPOS FALTANTES')
      except Exception as e:
                   print(e)
               

      return render(request,'Gestor/editarsegui.html',{'forma_persona':forma_persona,'seguimiento':seguimiento})
'''
aqui empieza el crud de actividades  
'''
  
@login_required
def calendario_activdades(request):
   actividad = AsignacionTarea.objects.all()
   return render(request,'Gestor/datercrud/actividadestareas.html',{'actividad':actividad})

@login_required
def guardar(request):
      actividad = AsignacionTareaForm()
      if request.method == 'POST':
            actividad = AsignacionTareaForm(request.POST)
            if actividad.is_valid():
                actividad = actividad.save()
                messages.add_message(request, messages.INFO, message='Se ha guardado con exito')
                return redirect('actividades')
            else:
                 messages.add_message(request, messages.ERROR, message='LLENE LOS CAMPOS FALTANTES')
                 actividad = AsignacionTareaForm()
      return render(request,'Gestor/datercrud/creartarea.html',{'actividad':actividad})

@login_required
def mostrarinfo(request,id):
    mostratinfodetalle=AsignacionTarea.objects.get(pk=id)
   
    return render(request,'Gestor/datercrud/mostrarmodalactividad.html',{'mostratinfodetalle':mostratinfodetalle})

@login_required
def editaractividad(request,id):
    mostratinfodetalle=AsignacionTarea.objects.get(pk=id)
    if request.method=='GET':
        actividad = AsignacionTareaForm(instance=mostratinfodetalle)
    else:
        actividad=AsignacionTareaForm(request.POST,instance=mostratinfodetalle)
        if actividad.is_valid():
            actividad = actividad.save()
            messages.add_message(request, messages.INFO, message='Se ha editado con exito')
            return redirect('actividades')
        else:
              messages.add_message(request, messages.ERROR, message='LLENE LOS CAMPOS FALTANTES')
    return render(request, 'Gestor/datercrud/editaractividadmodal.html',{'actividad':actividad,'mostrarinfodetalle':mostratinfodetalle})
# def obtener_gestor(_request):
#     gestor=AsignacionTarea.objects.all()
#     success =[]
#     for valor in gestor:
#         success.append(valor)

#     print(success)
  
        
#     return JsonResponse(success, safe=False)
    
@login_required
def actividadCrudDelete(request,id):

    asig=AsignacionTarea.objects.get(pk=id)
    print(asig)
    return render(request,'Gestor/datercrud/eliminaractividad.html',{'asig':asig})

@login_required
def ajax_eliminaractividad(request):
    id_actividad = request.POST.get('id')
    
    AsignacionTarea.objects.filter(id=id_actividad).update(estado_pendiente='0')
  
    messages.add_message(request, messages.INFO, message='Se elimino actividad')
    response={}
    return JsonResponse(response)

def entrada(request):
    return render(request,'entrada.html')
'''
envios de correos
'''
@login_required
def correo(request):
   try:
    if request.method=='POST':
        
            adjunto = request.FILES['adjunto']
            
            filestorege=FileSystemStorage(base_url='/media/',location='/media/')
            nombre=unidecode.unidecode(adjunto.name.replace(' ','_'))
            file = filestorege.save(nombre,adjunto)
            file_url=filestorege.url(file)
            filep=unidecode.unidecode(file_url)
            print(filep)
            para = request.POST["para"]
            asunto=request.POST.get("asunto")
            mensaje=request.POST["mensaje"]
            desde = settings.EMAIL_HOST_USER
            email = EmailMessage(asunto,mensaje,desde,to=[para])
        
        
            email.attach_file(file_url)
            email.fail_silenty=False
            email.send()
            messages.add_message(request, messages.SUCCESS, message='Se envio correo')
            return redirect('busqueda')
   except MultiValueDictKeyError as e:
            messages.add_message(request, messages.ERROR, message=f'Falta {e}')
            return redirect('correo')
        
   return render(request,'Gestor/correo.html')

'''
aqui empieza los reportes
'''

@login_required
def generar_report_caso(_request,id):
  caso = Casos.objects.get(pk=id)
  context = {'caso':caso,
             'hora_actual':datetime.datetime.now()}
  html = render_to_string("Gestor/reporte_por_caso.html", context)

  response = HttpResponse(content_type="application/pdf")
  response["Content-Disposition"] = "inline; report.pdf"

  
  HTML(string=html).write_pdf(response)
  return response
@login_required
def generar_report_graficas(_request):
    casos=Casos.objects.all()
    
    html = render_to_string("analista/graficas.html")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "inline; report.pdf"

        
    HTML(string=html).write_pdf(response)

   
  
    return response
@login_required
def generar_report_caso_general(_request):
  casos = Casos.objects.all()
  context = {'casos':casos,
             'hora_actual':datetime.datetime.now()}
  html = render_to_string("Gestor/generar_pdf_reporte_general.html", context)

  response = HttpResponse(content_type="application/pdf")
  response["Content-Disposition"] = "inline; report.pdf"

  
  HTML(string=html).write_pdf(response)

   
  
  return response



@login_required
def pagina_report(request):
    
    return render(request,'Gestor/selecciontiporeporte.html')

@login_required
def reportes_general_excel(request):
    casos = Casos.objects.all()
    wb=Workbook()
    bandera=True
    cont=1
 
    ws=wb.active 
    ws=wb.create_sheet('Hoja'+str(cont))
    ws['B1'].alignment=Alignment(horizontal='center',vertical='center')
    ws['B1'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['B1'].fill=PatternFill(start_color='66FFCC',end_color='66FFCC',fill_type='solid')
    ws['B1'].font=Font(name='Calibri',size=12,bold=True)
    ws['B1']='REPORTE GENERAL DE CASOS'
    ws.merge_cells('B1:AH1')
    ws.row_dimensions[1].height=25
    ws.column_dimensions['B'].width=20
    ws.column_dimensions['C'].width=20
    ws.column_dimensions['D'].width=20
    ws.column_dimensions['E'].width=20
    ws.column_dimensions['F'].width=20
    ws.column_dimensions['G'].width=20
    ws.column_dimensions['H'].width=20
    ws.column_dimensions['I'].width=20
    ws.column_dimensions['J'].width=20
    ws.column_dimensions['K'].width=20
    ws.column_dimensions['L'].width=20
    ws.column_dimensions['M'].width=20
    ws.column_dimensions['N'].width=20
    ws.column_dimensions['O'].width=20
    ws.column_dimensions['P'].width=20
    ws.column_dimensions['Q'].width=40
    ws.column_dimensions['R'].width=40
    ws.column_dimensions['S'].width=40
    ws.column_dimensions['T'].width=20
    ws.column_dimensions['U'].width=40
    ws.column_dimensions['V'].width=20
    ws.column_dimensions['W'].width=20
    ws.column_dimensions['X'].width=20
    ws.column_dimensions['Y'].width=20
    ws.column_dimensions['Z'].width=40
    ws.column_dimensions['AA'].width=40
    ws.column_dimensions['AB'].width=40
    ws.column_dimensions['AC'].width=40
    ws.column_dimensions['AD'].width=40
    ws.column_dimensions['AE'].width=40
    ws.column_dimensions['AF'].width=40
    ws.column_dimensions['AG'].width=30
    ws.column_dimensions['AH'].width=30
    ws['B3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['B3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['B3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['B3'].font=Font(name='Calibri',size=10,bold=True)
    ws['B3']='No caso'
    ws['C3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['C3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['C3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['C3'].font=Font(name='Calibri',size=10,bold=True)
    ws['C3']='Nombres'
    ws['D3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['D3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['D3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['D3'].font=Font(name='Calibri',size=10,bold=True)
    ws['D3']='Apellidos'
    ws['E3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['E3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['E3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['E3'].font=Font(name='Calibri',size=10,bold=True)
    ws['E3']='Identificacion'
    ws['F3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['F3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['F3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['F3'].font=Font(name='Calibri',size=10,bold=True)
    ws['F3']='Celular'
    ws['G3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['G3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['G3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['G3'].font=Font(name='Calibri',size=10,bold=True)
    ws['G3']='Correo'
    ws['H3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['H3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['H3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['H3'].font=Font(name='Calibri',size=10,bold=True)
    ws['H3']='Enfermedad'
    ws['I3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['I3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['I3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['I3'].font=Font(name='Calibri',size=10,bold=True)
    ws['I3']='Estado actual'
    ws['J3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['J3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['J3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['J3'].font=Font(name='Calibri',size=10,bold=True)
    ws['J3']='Gestor Farmaceutico'
    ws['K3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['K3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['K3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['K3'].font=Font(name='Calibri',size=10,bold=True)
    ws['K3']='Terapia'
    ws['L3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['L3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['L3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['L3'].font=Font(name='Calibri',size=10,bold=True)
    ws['L3']='Otra Terapia'
    ws['M3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['M3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['M3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['M3'].font=Font(name='Calibri',size=10,bold=True)
    ws['M3']='Tipo de requerimiento'
    ws['N3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['N3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['N3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['N3'].font=Font(name='Calibri',size=10,bold=True)
    ws['N3']='Clasificacion pbs'
    ws['O3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['O3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['O3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['O3'].font=Font(name='Calibri',size=10,bold=True)
    ws['O3']='Medico Tratante'
    ws['P3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['P3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['P3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['P3'].font=Font(name='Calibri',size=10,bold=True)
    ws['P3']='Seguna Barrera'
    ws['Q3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['Q3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['Q3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['Q3'].font=Font(name='Calibri',size=10,bold=True)
    ws['Q3']='Fecha de radicacion de la EPS'
    ws['R3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['R3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['R3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['R3'].font=Font(name='Calibri',size=10,bold=True)
    ws['R3']='Especialidad medica'       
    ws['S3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['S3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['S3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['S3'].font=Font(name='Calibri',size=10,bold=True)
    ws['S3']='Fecha de la formula médica' 
    ws['T3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['T3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['T3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['T3'].font=Font(name='Calibri',size=10,bold=True)
    ws['T3']='Fecha de autorizacion'
    ws['U3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['U3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['U3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['U3'].font=Font(name='Calibri',size=10,bold=True)
    ws['U3']='Fecha de radicacion formula medica'
    ws['V3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['V3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['V3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['V3'].font=Font(name='Calibri',size=10,bold=True)
    ws['V3']='Fecha de entrega'
    ws['W3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['W3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['W3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['W3'].font=Font(name='Calibri',size=10,bold=True)
    ws['W3']='Origen de solicitud'
    ws['X3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['X3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['X3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['X3'].font=Font(name='Calibri',size=10,bold=True)
    ws['X3']='IPS'
    ws['Y3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['Y3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['Y3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['Y3'].font=Font(name='Calibri',size=10,bold=True)
    ws['Y3']='Nombre del gestor'
    ws['Z3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['Z3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['Z3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['Z3'].font=Font(name='Calibri',size=10,bold=True)
    ws['Z3']='Fecha de registro del seguimiento'
    ws['AA3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['AA3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['AA3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['AA3'].font=Font(name='Calibri',size=10,bold=True)
    ws['AA3']='Descripcion del seguimiento'
    ws['AB3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['AB3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['AB3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['AB3'].font=Font(name='Calibri',size=10,bold=True)
    ws['AB3']='Fecha de atencion de EPS' 
    ws['AC3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['AC3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['AC3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['AC3'].font=Font(name='Calibri',size=10,bold=True)
    ws['AC3']='Fecha de estado abierta' 
    ws['AD3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['AD3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['AD3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['AD3'].font=Font(name='Calibri',size=10,bold=True)
    ws['AD3']='Fecha de estado en proceso' 
    ws['AE3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['AE3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['AE3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['AE3'].font=Font(name='Calibri',size=10,bold=True)
    ws['AE3']='Fecha de estado finalizado'
    ws['AF3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['AF3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['AF3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['AF3'].font=Font(name='Calibri',size=10,bold=True)
    ws['AF3']='No radicado' 
    ws['AG3'].alignment=Alignment(horizontal='center',vertical='center')
    ws['AG3'].border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
    ws['AG3'].fill=PatternFill(start_color='66CFCC',end_color='66CFCC',fill_type='solid')
    ws['AG3'].font=Font(name='Calibri',size=10,bold=True)
    ws['AG3']='Descripcion del caso'
    cantidad=4
    for caso in casos:
      
        ws.cell(row=cantidad,column=2).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=2).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=2).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=2).value=caso.id_caso
        ws.cell(row=cantidad,column=3).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=3).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=3).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=3).value=caso.id_usuario.primer_nombre + caso.id_usuario.segundo_nombre
        ws.cell(row=cantidad,column=4).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=4).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=4).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=4).value=caso.id_usuario.primer_apellido + caso.id_usuario.segundo_apellido
        ws.cell(row=cantidad,column=5).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=5).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=5).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=5).value=caso.id_usuario.id_cedula
        ws.cell(row=cantidad,column=6).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=6).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=6).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=6).value=caso.id_usuario.celular
        ws.cell(row=cantidad,column=7).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=7).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=7).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=7).value=caso.id_usuario.login_id.email
        ws.cell(row=cantidad,column=8).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=8).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=8).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=8).value=caso.enfermedad.nombreenfermedad
        ws.cell(row=cantidad,column=9).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=9).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=9).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=9).value=caso.estado.nombreestado
        ws.cell(row=cantidad,column=10).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=10).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=10).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=10).value=caso.id_comple_info.gestor_farma.nombrefarmacia
        ws.cell(row=cantidad,column=11).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=11).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=11).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=11).value=caso.id_comple_info.terapia.nombreterapia
        ws.cell(row=cantidad,column=12).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=12).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=12).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=12).value=caso.id_comple_info.otra_terapia
        ws.cell(row=cantidad,column=13).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=13).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=13).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=13).value=caso.id_comple_info.tipo_req.nombrerequerimiento
        ws.cell(row=cantidad,column=14).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=14).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=14).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=14).value=caso.id_comple_info.clasificacion_pbs.nombrepbs
        ws.cell(row=cantidad,column=15).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=15).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=15).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=15).value=caso.id_comple_info.medico_trat 
        ws.cell(row=cantidad,column=16).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=16).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=16).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=16).value=caso.id_comple_info.segunda_barrera
        ws.cell(row=cantidad,column=17).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=17).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=17).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=17).value=caso.id_comple_info.fech_rad_for_eps
        ws.cell(row=cantidad,column=18).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=18).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=18).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=18).value=caso.id_comple_info.especialidad_med.nombre
        
        ws.cell(row=cantidad,column=19).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=19).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=19).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=19).value=caso.id_comple_info.fecha_for_medi
        ws.cell(row=cantidad,column=20).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=20).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=20).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=20).value=caso.id_comple_info.fecha_aut
        ws.cell(row=cantidad,column=21).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=21).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=21).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=21).value=caso.id_comple_info.fech_rad_aut_farm
        ws.cell(row=cantidad,column=22).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=22).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=22).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=22).value=caso.id_comple_info.fecha_entrega
        ws.cell(row=cantidad,column=23).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=23).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=23).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=23).value=caso.id_comple_info.origen_soli
        ws.cell(row=cantidad,column=24).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=24).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=24).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=24).value=caso.id_comple_info.ips_id_terapia.nombre
        ws.cell(row=cantidad,column=25).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=25).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=25).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=25).value=caso.id_gest.id_datos_us.primer_nombre
        ws.cell(row=cantidad,column=26).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=26).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=26).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=26).value=caso.id_seguimiento.fecharegistro
        ws.cell(row=cantidad,column=27).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=27).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=27).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=27).value=caso.id_seguimiento.descripcion
        ws.cell(row=cantidad,column=28).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=28).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=28).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=28).value=caso.fecharesgistrocaso
        ws.cell(row=cantidad,column=28).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=28).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=28).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=28).value=caso.fechaatencioneps
        ws.cell(row=cantidad,column=29).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=29).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=29).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=29).value=caso.fechaatenabierto
        ws.cell(row=cantidad,column=30).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=30).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=30).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=30).value=caso.fechaatenproceso
        ws.cell(row=cantidad,column=31).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=31).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=31).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=31).value=caso.fechaatenfinalizado
        ws.cell(row=cantidad,column=32).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=32).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=32).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=32).value=caso.numeroradicado
        ws.cell(row=cantidad,column=33).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(row=cantidad,column=33).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
        ws.cell(row=cantidad,column=33).font=Font(name='Calibri',size=10,bold=True)
        ws.cell(row=cantidad,column=33).value=caso.descripcioncaso
        
        cantidad+=1
        
     
        
    nombre_del_archivo='ReporteCaso.xlsx'
    response=HttpResponse(content_type='application/ms-excel')
    contenido='attachment;filename = {0}'.format(nombre_del_archivo)
    response['Content-Disposition']=contenido
    wb.save(response)
    return response
'''
aqui empieza las graficas 
'''
@login_required
def vista_graficas(request):
    casos=Casos.objects.all()
    proceso=[]
    labelsproceso=[]
    finalizado=[]
    labelsfinalizado=[]
    abierto=[]
    labelsabierto=[]
    estados=[]
    labelsestados=[]
    epss=[]
    labelseps=[]
    reg=[]
    labelsreg=[]
    regcont=[]
    labelsregcont=[]
    sub=[]
    labelssub=[]
    depart=[]
    labelsdep=[]
    muni=[]
    labelsmuni=[]
    generos=[]
    labelsge=[]
    enfer=[]
    labelsenfe=[]
    medic=[]
    labelsmedic=[]
    barr=[]
    labelsbarr=[]
    queryabierto= Casos.objects.values('estado__nombreestado').filter(estado__nombreestado='abierto').annotate(cant_estado=Count('estado'))
    queryproceso= Casos.objects.values('estado__nombreestado').filter(estado__nombreestado='proceso').annotate(cant_estado=Count('estado'))
    queryfinalizado= Casos.objects.values('estado__nombreestado').filter(estado__nombreestado='finalizado').annotate(cant_estado=Count('estado'))
    queryestados=Casos.objects.values('estado__nombreestado').annotate(cant_estado=Count('estado'))
    eps = Casos.objects.values('id_usuario__id_eps__nombre').annotate(cant_eps=Count('id_usuario__id_eps'))
    regimen=Casos.objects.values('id_usuario__idtiporegimen__nombreregimen').annotate(cant_reg=Count('id_usuario__idtiporegimen__nombreregimen'))
    regimencont=Casos.objects.values('id_usuario__idtiporegimen__nombreregimen').filter(id_usuario__idtiporegimen__nombreregimen='Contributivo').annotate(cant_cont=Count('id_usuario__idtiporegimen'))
    regimensub=Casos.objects.values('id_usuario__idtiporegimen__nombreregimen').filter(id_usuario__idtiporegimen__nombreregimen='subsidiado').annotate(cant_sub=Count('id_usuario__idtiporegimen'))
    departamento=Casos.objects.values('id_usuario__cod_dep__nombre').annotate(cant_dep=Count('id_usuario__cod_dep'))
    municipio = Casos.objects.values('id_usuario__cod_muni__nombremunicipio').annotate(cant_mun=Count('id_usuario__cod_muni'))
    genero = Casos.objects.values('id_usuario__idgenero__nombregenero').annotate(cant_genero=Count('id_usuario__idgenero'))
    enfermedad=Casos.objects.values('enfermedad__nombreenfermedad').annotate(enfermedad=Count('enfermedad'))
    medicamentos=Casos.objects.values('id_comple_info__clasificacion_pbs__nombrepbs').annotate(medicamentos=Count('id_comple_info__clasificacion_pbs'))
    barrera=Casos.objects.values('id_barrera__nombre').annotate(barreras=Count('id_barrera'))
    
    for x in barrera:
        barr.append(x['id_barrera__nombre'])
        labelsbarr.append(x['barreras'])
    for x in medicamentos:
        medic.append(x['id_comple_info__clasificacion_pbs__nombrepbs'])
        labelsmedic.append(x['medicamentos'])
    for x in enfermedad:
        enfer.append(x['enfermedad__nombreenfermedad'])
        labelsenfe.append(x['enfermedad'])
    for x in genero:
        generos.append(x['id_usuario__idgenero__nombregenero'])
        labelsge.append(x['cant_genero'])
    for x in municipio:
        muni.append(x['id_usuario__cod_muni__nombremunicipio'])
        labelsmuni.append(x['cant_mun'])
    for z in departamento:
        depart.append(z['id_usuario__cod_dep__nombre'])
        labelsdep.append(z['cant_dep'])
    print(depart)
    for x in regimensub:
        sub.append(x['id_usuario__idtiporegimen__nombreregimen'])
        labelssub.append(x['cant_sub'])
    for x in regimencont:
        regcont.append(x['id_usuario__idtiporegimen__nombreregimen'])
        labelsregcont.append(x['cant_cont'])
    for r in regimen:
        reg.append(r['id_usuario__idtiporegimen__nombreregimen'])
        labelsreg.append(r['cant_reg'])
    for e in eps:
        epss.append(e['id_usuario__id_eps__nombre'])
        labelseps.append(e['cant_eps'])
    for abiertos in queryabierto:
        abierto.append(abiertos['estado__nombreestado'])
        labelsabierto.append(abiertos['cant_estado'])
    for abiertos in queryproceso:
        proceso.append(abiertos['estado__nombreestado'])
        labelsproceso.append(abiertos['cant_estado'])
    for abiertos in queryfinalizado:
        finalizado.append(abiertos['estado__nombreestado'])
        labelsfinalizado.append(abiertos['cant_estado'])
    for nestados in queryestados:
        estados.append(nestados['estado__nombreestado'])
        labelsestados.append(nestados['cant_estado'])
    print(epss)
    return render(request,'analista/graficas.html',
                  {
                   'labelsestados':labelsestados,
                   'estados':estados,
                   'labelsabierto':labelsabierto,
                   'abierto':abierto,
                   'labelsfinalizado':labelsfinalizado,
                   'finalizado':finalizado,
                   'labelsproceso':labelsproceso,
                   'proceso':proceso,
                   'labelseps':labelseps,
                   'EPS':epss,
                   'labelsreg':labelsreg,
                   'reg':reg,
                   'casos':casos,
                   'contributivo':regcont,
                   'labelscont':labelsregcont,
                   'subsidiado':labelssub,
                   'departamento':depart,
                   'labelsdep':labelsdep,
                   'municipio':muni,
                   'labelsmuni':labelsmuni,
                   'genero':generos,
                   'labelsge':labelsge,
                   'enfermedad':enfer,
                   'labelsenfer':labelsenfe,
                   'medic':medic,
                   'labelsmedic':labelsmedic,
                   'barrera':barr,
                   'labelsbarr':labelsbarr,
                   
                                                    })

class TimeStampDiff(Func):
    class PrettyStringFormatting(dict):
        def __missing__(self, key):
            return '%(' + key + ')s'

    def __init__(self, *expressions, **extra):
        unit = extra.pop('unit', 'day')
        self.template = self.template % self.PrettyStringFormatting({"unit": unit})
        super().__init__(*expressions, **extra)

    function = 'TIMESTAMPDIFF'
    template = "%(function)s(%(unit)s, %(expressions)s)"
@login_required
def get_data(request):

    lista=[]
    casosregistrados= Casos.objects.values('id_caso').aggregate(total=Count('id_caso'))
    cantidadusuarios=DatosUsuario.objects.values('id_cedula').aggregate(total=Count('id_cedula'))
    casosresueltos = Casos.objects.values('id_caso').filter(estado__nombreestado='finalizado').aggregate(total=Count('id_caso'))
    casos_activos_proceso=Casos.objects.values('estado').filter(estado__nombreestado='proceso').aggregate(valor=Count('id_caso'))
    casos_activos_abiertos=Casos.objects.values('estado').filter(estado__nombreestado='abierto').aggregate(valor=Count('id_caso'))
    cantidad_activos=casos_activos_proceso['valor'] + casos_activos_abiertos['valor']
    fechafinal = Casos.objects.values('id_caso').annotate(cantida_dias=Sum(TimeStampDiff(F('fechaatenfinalizado') ,F('fechaatenabierto'),output_field=IntegerField())))
    list_new=[]
    sum=0
    for x in fechafinal:
        lista.append(x['cantida_dias'])
    
    for i in lista:
        if i !=None:
            list_new.append(i)

    for x in list_new:
          sum+=x
    
    
      

    promedio_respuesta_dias=round((sum/30)*100)
    
    data={
      
        'casosregistradostotal':casosregistrados,
        'cantusers':cantidadusuarios,
        'resueltos':casosresueltos,
        'activos':cantidad_activos,
        'promedio_respuesta_dias':promedio_respuesta_dias
    }

    return JsonResponse(data)

@login_required
def indicadores_gestion(request):
    
    return render(request,'Gestor/indicadores.html')
  
    